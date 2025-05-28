
from flask import Flask, render_template_string, request, redirect, url_for
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
EXCEL_PATH = "tickets.xlsx"

if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    ws.append(["Data", "Projeto", "Solicitante", "Solicitação", "Obs", "Status"])
    wb.save(EXCEL_PATH)

HTML_FORM = """<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <title>Abertura de Chamado</title>
    <style>
        body { font-family: Arial, sans-serif; background: #f0f0f0; padding: 30px; }
        form { background: white; padding: 20px; border-radius: 8px; max-width: 500px; margin: auto; }
        label { display: block; margin-top: 15px; }
        input, textarea { width: 100%; padding: 8px; margin-top: 5px; }
        button { margin-top: 20px; padding: 10px 15px; }
        a { display: block; margin-top: 20px; text-align: center; }
    </style>
</head>
<body>
    <h2 style="text-align:center;">Formulário de Abertura de Chamado</h2>
    <form method="post">
        <label>Projeto:<input name="projeto" required></label>
        <label>Solicitante:<input name="solicitante" required></label>
        <label>Solicitação:<textarea name="solicitacao" rows="4" required></textarea></label>
        <label>Observações:<textarea name="obs" rows="2"></textarea></label>
        <label>Status:<input name="status" value="Aberto" readonly></label>
        <button type="submit">Enviar Chamado</button>
    </form>
    <a href="/chamados">Ver Chamados</a>
</body>
</html>"""

HTML_CHAMADOS = """<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <title>Chamados Recebidos</title>
    <style>
        body { font-family: Arial, sans-serif; background: #f0f0f0; padding: 30px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; background: white; }
        th, td { border: 1px solid #ccc; padding: 8px; text-align: left; }
        th { background-color: #eee; }
        a { display: block; margin-bottom: 20px; text-align: center; }
    </style>
</head>
<body>
    <a href="/">← Voltar para o formulário</a>
    <h2 style="text-align:center;">Chamados Recebidos</h2>
    <table>
        <tr>
            <th>Data</th>
            <th>Projeto</th>
            <th>Solicitante</th>
            <th>Solicitação</th>
            <th>Obs</th>
            <th>Status</th>
        </tr>
        {% for row in chamados %}
        <tr>
            {% for cell in row %}
            <td>{{ cell }}</td>
            {% endfor %}
        </tr>
        {% endfor %}
    </table>
</body>
</html>"""

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        data = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        projeto = request.form["projeto"]
        solicitante = request.form["solicitante"]
        solicitacao = request.form["solicitacao"]
        obs = request.form.get("obs", "")
        status = request.form.get("status", "Aberto")

        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append([data, projeto, solicitante, solicitacao, obs, status])
        wb.save(EXCEL_PATH)
        return redirect(url_for("form"))

    return render_template_string(HTML_FORM)

@app.route("/chamados")
def chamados():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    chamados_data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
    return render_template_string(HTML_CHAMADOS, chamados=chamados_data)

if __name__ == "__main__":
    app.run(debug=True)
